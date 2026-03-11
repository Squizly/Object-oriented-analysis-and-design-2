import openpyxl
import os
from models import University, Student, Lecturer

class DataManager:
    FILE_NAME = "data_to_import.xlsx"

    @staticmethod
    def load_data(university: University):
        university.clear()
        if not os.path.exists(DataManager.FILE_NAME):
            raise FileNotFoundError(f"Файл {DataManager.FILE_NAME} не найден!")

        wb = openpyxl.load_workbook(DataManager.FILE_NAME)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            
            p_type, name, age, faculty, email = row[0], row[1], row[2], row[3], row[4]
            student_id, course, gpa = row[5], row[6], row[7]
            post, exp = row[8], row[9]

            if p_type == "Student":
                university.addPers(Student(name, int(age or 0), faculty, email, str(student_id), int(course or 0), float(gpa or 0.0)))
            elif p_type == "Lecturer":
                university.addPers(Lecturer(name, int(age or 0), faculty, email, str(post), int(exp or 0)))

    @staticmethod
    def save_data(university: University):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Type", "FullName", "Age", "Faculty", "Email", "StudentID", "Course", "GPA", "Post", "Experience"])

        for p in university.persons:
            if isinstance(p, Student):
                ws.append(["Student", p.fullName, p.age, p.faculty, p.email, p.studentID, p.course, p.GPA, "", ""])
            elif isinstance(p, Lecturer):
                ws.append(["Lecturer", p.fullName, p.age, p.faculty, p.email, "", "", "", p.post, p.experience])

        wb.save(DataManager.FILE_NAME)